"""Import MealRegistration (and optionally MealPreference) rows from the
KB madtilmeldinger format.

Accepts:
  - A local CSV file (single sheet)
  - A local XLSX file (all week-numbered sheets imported)
  - A Google Sheets spreadsheet ID or URL (auto-downloads all CSV sheets)
  - A Google Drive XLSX file ID (auto-downloaded via Drive)

Usage examples::

    # Import all weeks from a local XLSX file
    python manage.py import_madtilmeldinger /path/to/MadRegnskab_jan_2026.xlsx --year 2026

    # Import from a Google Drive XLSX file ID
    python manage.py import_madtilmeldinger 1Kpgg2LXXwWk0c4KgbphXk4roe20mSwaZ --year 2026

    # Import all sheets from the Google Sheets spreadsheet (weeks 22-25)
    python manage.py import_madtilmeldinger 1cT7BXrmP9IG60skxjkaV2peay4RF99fl --year 2026

    # Also import 'Fast tilmelding' sheet as MealPreference defaults
    python manage.py import_madtilmeldinger 1cT7... --year 2026 --import-preferences

    # Dry-run (no changes committed)
    python manage.py import_madtilmeldinger 1cT7... --year 2026 --dry-run

    # Clear all existing MealRegistration data first (test data cleanup)
    python manage.py import_madtilmeldinger 1cT7... --year 2026 --clear

Spreadsheet layout (rows, 0-indexed):
  0:   header comment ("V=Vegetar W=Weganer KF= Kød/Fisk …")
  1:   optional holiday notes, e.g. ",Pinsedag,,,…" (col corresponds to day)
  2:   day category headers ("Uge/Fast, Mandag (Vegetar), …, Torsdag (Vegetar), …")
  3:   week number (digit) or "tilmelding" for the Fast-tilmelding (preferences) sheet
  4:   column headers: Husnr, V, W, 1-11, V, W, 1-11, V, W, KF, 1-11, V, W, 1-11, Husnr, Pris
  5+:  one row per house; rows stop when col-0 is not a plain integer

Column mapping (0-indexed):
  Monday:    V=1, W=2, children=3
  Tuesday:   V=4, W=5, children=6
  Wednesday: V=7, W=8, KF(meat)=9, children=10
  Thursday:  V=11, W=12, children=13

V + W → adults_veg  |  KF → adults_meat (Wednesday only)  |  1-11 → children_count

Sheets without "Torsdag" in the day-headers row are skipped (old 3-day format, pre-2026).
"""

import csv
import datetime
import io
import re
import urllib.request
from datetime import date, timedelta
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.food.models import ClosedFoodDay, DayOfWeek, MealPreference, MealRegistration
from apps.houses.models import House

# Column positions per day (V cols + W col; V and W both count as adults_veg)
_DAY_COLS: dict[DayOfWeek, dict] = {
    DayOfWeek.MONDAY: {"veg": [1, 2], "meat": None, "children": 3},
    DayOfWeek.TUESDAY: {"veg": [4, 5], "meat": None, "children": 6},
    DayOfWeek.WEDNESDAY: {"veg": [7, 8], "meat": 9, "children": 10},
    DayOfWeek.THURSDAY: {"veg": [11, 12], "meat": None, "children": 13},
}

# Row-1 column position where a holiday note appears for each day
_HOLIDAY_NOTE_COL: dict[DayOfWeek, int] = {
    DayOfWeek.MONDAY: 1,
    DayOfWeek.TUESDAY: 4,
    DayOfWeek.WEDNESDAY: 7,
    DayOfWeek.THURSDAY: 11,
}


def _xlsx_cell(v: object) -> str:
    """Normalise an XLSX cell value to a plain string (matches CSV field behaviour)."""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        # Excel misreads "1-11" (children header) as Nov 1 — treat as empty string.
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)


def _to_int(cell: str) -> int:
    try:
        return int(str(cell).strip())
    except (ValueError, TypeError):
        return 0


def _monday_of_iso_week(year: int, week: int) -> date:
    return date.fromisocalendar(year, week, 1)


def _get_col(row: list[str], idx: int) -> str:
    return row[idx] if len(row) > idx else ""


def _extract_day_values(row: list[str], day: DayOfWeek) -> tuple[int, int, int]:
    """Return (adults_veg, adults_meat, children_count) for a day from a house row."""
    cols = _DAY_COLS[day]
    adults_veg = sum(_to_int(_get_col(row, c)) for c in cols["veg"])
    adults_meat = _to_int(_get_col(row, cols["meat"])) if cols["meat"] is not None else 0
    children = _to_int(_get_col(row, cols["children"]))
    return adults_veg, adults_meat, children


def _parse_sheet(rows: list[list[str]]) -> dict:
    """Parse a sheet's rows and return structured data.

    Returns a dict with keys:
      skip (bool), reason (str)
      is_preferences (bool), week_number (int|None)
      holiday_notes (dict[DayOfWeek, str])
      house_data (list[list[str]])
    """
    if len(rows) < 6:
        return {"skip": True, "reason": "too few rows"}

    # Validate: row 4 must have "Husnr" in col 0
    if _get_col(rows[4], 0).strip() != "Husnr":
        return {"skip": True, "reason": f"col 0 of row 4 is {_get_col(rows[4], 0)!r}, not 'Husnr'"}

    # Skip old 3-day format (pre-2026) which lacks a Thursday column
    if not any("Torsdag" in str(_get_col(rows[2], c)) for c in range(len(rows[2]))):
        return {"skip": True, "reason": "old 3-day format (no Thursday column)"}

    # Holiday notes in row 1
    holiday_notes: dict[DayOfWeek, str] = {}
    for day, col in _HOLIDAY_NOTE_COL.items():
        note = _get_col(rows[1], col).strip()
        if note:
            holiday_notes[day] = note

    # Week number or preferences marker (row 3, col 0)
    week_cell = _get_col(rows[3], 0).strip()
    is_preferences = not week_cell.isdigit()
    week_number = None if is_preferences else int(week_cell)

    # House rows start at row 5; stop when col 0 is not a plain integer
    house_data = []
    for row in rows[5:]:
        husnr = _get_col(row, 0).strip()
        if not husnr or not re.match(r"^\d+$", husnr):
            break
        house_data.append(row)

    return {
        "skip": False,
        "is_preferences": is_preferences,
        "week_number": week_number,
        "holiday_notes": holiday_notes,
        "house_data": house_data,
    }


class Command(BaseCommand):
    help = (
        "Import MealRegistration (and optionally MealPreference) records "
        "from the KB madtilmeldinger Google Sheets CSV format."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "source",
            help=(
                "Path to a local CSV file, or a Google Sheets spreadsheet ID / URL "
                "(public spreadsheet; all sheets auto-downloaded)."
            ),
        )
        parser.add_argument(
            "--year",
            type=int,
            required=True,
            help="Calendar year for the ISO week numbers in the spreadsheet (e.g. 2026).",
        )
        parser.add_argument(
            "--import-preferences",
            action="store_true",
            help="Also import the 'Fast tilmelding' sheet as MealPreference records.",
        )
        parser.add_argument(
            "--import-closed-days",
            action="store_true",
            help=(
                "Create ClosedFoodDay records for days with holiday notes "
                "(e.g. 'Pinsedag'). Uses get_or_create — safe to run multiple times."
            ),
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help=(
                "Delete ALL existing MealRegistration records before importing. "
                "With --import-preferences also clears MealPreference. "
                "Useful when all existing data is test data."
            ),
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report without committing any changes.",
        )

    def handle(
        self,
        *args,
        source: str,
        year: int,
        import_preferences: bool,
        import_closed_days: bool,
        clear: bool,
        dry_run: bool,
        **options,
    ):
        houses: dict[str, House] = {h.slug: h for h in House.objects.all() if h.slug.isdigit()}
        if not houses:
            raise CommandError("No houses found in the database. Run seed_houses first.")

        sheets = self._load_sheets(source)

        created_regs = updated_regs = 0
        created_prefs = updated_prefs = 0
        created_closed = 0
        missing_houses: list[str] = []

        with transaction.atomic():
            if clear:
                reg_count, _ = MealRegistration.objects.all().delete()
                self.stdout.write(f"Cleared {reg_count} MealRegistration records.")
                if import_preferences:
                    pref_count, _ = MealPreference.objects.all().delete()
                    self.stdout.write(f"Cleared {pref_count} MealPreference records.")

            for sheet_name, rows in sheets:
                parsed = _parse_sheet(rows)
                if parsed["skip"]:
                    self.stdout.write(f"  Skipping '{sheet_name}': {parsed['reason']}")
                    continue

                if parsed["is_preferences"]:
                    if not import_preferences:
                        self.stdout.write(
                            f"  Skipping 'Fast tilmelding' sheet '{sheet_name}' "
                            f"(pass --import-preferences to include it)."
                        )
                        continue
                    c, u = self._import_preferences(parsed, houses, missing_houses)
                    created_prefs += c
                    updated_prefs += u
                    self.stdout.write(
                        f"  '{sheet_name}' (fast tilmelding): "
                        f"MealPreference created={c} updated={u}"
                    )
                else:
                    week = parsed["week_number"]
                    monday = _monday_of_iso_week(year, week)

                    c, u = self._import_registrations(parsed, houses, monday, missing_houses)
                    created_regs += c
                    updated_regs += u

                    holiday_str = ""
                    if parsed["holiday_notes"]:
                        notes = ", ".join(
                            f"{d.label}: {n}" for d, n in parsed["holiday_notes"].items()
                        )
                        holiday_str = f" [holidays: {notes}]"

                    self.stdout.write(
                        f"  '{sheet_name}' (uge {week}, from {monday}): "
                        f"MealRegistration created={c} updated={u}{holiday_str}"
                    )

                    if import_closed_days:
                        for day, note in parsed["holiday_notes"].items():
                            day_date = monday + timedelta(days=day.value)
                            _, was_created = ClosedFoodDay.objects.get_or_create(
                                date=day_date,
                                defaults={"reason": note},
                            )
                            if was_created:
                                created_closed += 1
                                self.stdout.write(f"    ClosedFoodDay: {day_date} ({note})")

            if dry_run:
                transaction.set_rollback(True)

        prefix = "(dry-run) " if dry_run else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"\n{prefix}Done — "
                f"Registrations: created={created_regs} updated={updated_regs}  "
                f"Preferences: created={created_prefs} updated={updated_prefs}  "
                f"ClosedFoodDays: created={created_closed}"
            )
        )
        if missing_houses:
            unique = sorted(set(missing_houses), key=lambda x: int(x))
            self.stdout.write(self.style.WARNING(f"No House found for husnr: {', '.join(unique)}"))

    # ------------------------------------------------------------------
    # DB writes
    # ------------------------------------------------------------------

    def _import_registrations(
        self,
        parsed: dict,
        houses: dict[str, House],
        monday: date,
        missing_houses: list[str],
    ) -> tuple[int, int]:
        created = updated = 0
        for row in parsed["house_data"]:
            husnr = _get_col(row, 0).strip()
            house = houses.get(husnr)
            if house is None:
                missing_houses.append(husnr)
                continue

            for day in DayOfWeek:
                veg, meat, children = _extract_day_values(row, day)
                if veg == 0 and meat == 0 and children == 0:
                    continue  # House did not register for this day

                day_date = monday + timedelta(days=day.value)
                _, was_created = MealRegistration.objects.update_or_create(
                    house=house,
                    date=day_date,
                    defaults={
                        "adults_veg": veg,
                        "adults_meat": meat,
                        "children_count": children,
                        "is_active": True,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        return created, updated

    def _import_preferences(
        self,
        parsed: dict,
        houses: dict[str, House],
        missing_houses: list[str],
    ) -> tuple[int, int]:
        created = updated = 0
        for row in parsed["house_data"]:
            husnr = _get_col(row, 0).strip()
            house = houses.get(husnr)
            if house is None:
                missing_houses.append(husnr)
                continue

            for day in DayOfWeek:
                veg, meat, children = _extract_day_values(row, day)
                if veg == 0 and meat == 0 and children == 0:
                    continue  # No default for this day

                _, was_created = MealPreference.objects.update_or_create(
                    house=house,
                    day_of_week=day.value,
                    defaults={
                        "adults_veg": veg,
                        "adults_meat": meat,
                        "children_count": children,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        return created, updated

    # ------------------------------------------------------------------
    # Source loading
    # ------------------------------------------------------------------

    def _load_sheets(self, source: str) -> list[tuple[str, list[list[str]]]]:
        path = Path(source).expanduser()
        if path.exists():
            if path.suffix.lower() == ".xlsx":
                return self._read_xlsx(path)
            return [(path.name, self._read_csv(path.read_text(encoding="utf-8-sig")))]

        drive_id = self._parse_drive_id(source)
        if not drive_id:
            raise CommandError(
                f"'{source}' is not a valid file path or Google Drive ID/URL. "
                f"IDs look like: 1cT7BXrmP9IG60skxjkaV2peay4RF99fl"
            )

        # Try as Google Sheets (CSV export) first; fall back to Drive XLSX download.
        sheets_html_url = f"https://docs.google.com/spreadsheets/d/{drive_id}/htmlview"
        try:
            with urllib.request.urlopen(sheets_html_url, timeout=15) as resp:
                html = resp.read().decode("utf-8", errors="ignore")
            gids = list(dict.fromkeys(re.findall(r"gid=([0-9]+)", html)))
            if gids:
                return self._download_sheets_csv(drive_id, gids)
        except Exception:
            pass

        # Fall back: download as XLSX from Google Drive
        return self._download_drive_xlsx(drive_id)

    @staticmethod
    def _parse_drive_id(source: str) -> str | None:
        """Extract a Drive/Sheets file ID from a URL or treat source as a bare ID."""
        for pattern in (r"/spreadsheets/d/([a-zA-Z0-9_-]+)", r"/file/d/([a-zA-Z0-9_-]+)"):
            m = re.search(pattern, source)
            if m:
                return m.group(1)
        if re.match(r"^[a-zA-Z0-9_-]{20,}$", source):
            return source
        return None

    def _download_sheets_csv(
        self, spreadsheet_id: str, gids: list[str]
    ) -> list[tuple[str, list[list[str]]]]:
        self.stdout.write(
            f"Google Sheets {spreadsheet_id}: {len(gids)} sheets (gids: {', '.join(gids)})"
        )
        results = []
        for gid in gids:
            export_url = (
                f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
                f"/export?format=csv&gid={gid}"
            )
            try:
                with urllib.request.urlopen(export_url, timeout=15) as resp:
                    content = resp.read().decode("utf-8-sig", errors="ignore")
            except Exception as exc:
                self.stdout.write(self.style.WARNING(f"  Failed gid={gid}: {exc}"))
                continue
            results.append((f"gid={gid}", self._read_csv(content)))
        return results

    def _download_drive_xlsx(self, file_id: str) -> list[tuple[str, list[list[str]]]]:
        """Download a Google Drive XLSX file by ID and return its week sheets."""
        import tempfile

        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        self.stdout.write(f"Drive XLSX {file_id}: downloading…")
        try:
            with urllib.request.urlopen(download_url, timeout=30) as resp:
                data = resp.read()
        except Exception as exc:
            raise CommandError(f"Failed to download XLSX from Drive ({file_id}): {exc}") from exc

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)

        try:
            return self._read_xlsx(tmp_path)
        finally:
            tmp_path.unlink(missing_ok=True)

    def _read_xlsx(self, path: Path) -> list[tuple[str, list[list[str]]]]:
        """Read all week-numbered sheets from an XLSX file."""
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required to read XLSX files: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(path, data_only=True)
        self.stdout.write(f"XLSX {path.name}: sheets = {wb.sheetnames}")

        results = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[_xlsx_cell(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
            # Trim trailing all-empty rows
            while rows and all(v == "" for v in rows[-1]):
                rows.pop()
            if rows:
                results.append((sheet_name, rows))

        return results

    @staticmethod
    def _read_csv(content: str) -> list[list[str]]:
        return list(csv.reader(io.StringIO(content)))
