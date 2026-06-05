"""
Service for fetching and parsing recipe spreadsheets from Google Drive.

Each week's Drive folder (DriveMenuCache.drive_folder_id) contains ONE
spreadsheet whose worksheet tabs are named like ``Ma1, Ma2, Ti1, Ti2,
On1, On2, To1, To2`` (Ma=Mandag, Ti=Tirsdag, On=Onsdag, To=Torsdag;
trailing digit is the dish index). The dish name lives in cell **C1**
(columns A and B are hidden ingredient-quantity columns).

The parsed result is cached on the DriveMenuCache row in ``recipe_sheets``
(a list of {code, day, index, name, weekday, url, ingredients, steps, _v}
dicts) along with the spreadsheet's Drive file id in ``recipe_file_id``.
``ingredients`` is a list of {amount, unit, name, comment} and ``steps`` a
list of Fremgangsmåde paragraphs, so the app can render the full recipe
in-app. Each entry carries a ``_v`` schema version so a parser change
automatically invalidates previously-cached rows on the next read.
"""

import datetime
import io
import logging
import re

from django.conf import settings
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

logger = logging.getLogger(__name__)

# Sheet tab name pattern, e.g. "Ma1", "To2" (case-insensitive)
SHEET_PATTERN = re.compile(r"^(Ma|Ti|On|To)(\d+)$", re.IGNORECASE)

# Day prefix -> weekday index (Mon=0 ... Thu=3)
DAY_PREFIX_TO_INDEX = {
    "ma": 0,
    "ti": 1,
    "on": 2,
    "to": 3,
}

# Display labels per weekday index. Derived from the tab code (the source of
# truth) rather than the sheet's F1 cell, which is sometimes a stale copy-paste
# (e.g. a "To3" tab whose F1 still reads "Onsdag").
DANISH_WEEKDAYS = ["Mandag", "Tirsdag", "Onsdag", "Torsdag"]

GSHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Recipe-sheet layout (1-indexed columns, openpyxl/Sheets-API convention).
# Columns A and B are alternate quantity scalings; the human-readable data the
# kitchen reads lives in C–F (per the menu template):
#   C1            -> dish name        F1 -> weekday label ("Mandag" …)
#   <header> row  -> "Ingrediens" in col E marks the ingredient table header
#   ingredient    -> C amount, D unit, E name, F comment
#   "Fremgangsmåde" header row, then each step is a cell merged across C:F
#   (the text sits in column C; merged continuation rows are blank -> skipped)
DISH_NAME_CELL = "C1"
DISH_NAME_COL = 3  # 1-indexed column for openpyxl / dish name
DISH_NAME_ROW = 1
COL_UNIT = 4  # D
COL_INGREDIENT = 5  # E
COL_COMMENT = 6  # F
COL_STEP = 3  # C — merged Fremgangsmåde step text
# Preferred amount column, then fallbacks. Most sheets fill C; a few only fill
# B (or A), so we pick the first of these that has any value in the table.
AMOUNT_COL_PREFERENCE = (3, 2, 1)  # C, then B, then A
# Markers (lower-cased) that delimit the two sections of a recipe sheet.
INGREDIENT_HEADER = "ingrediens"
METHOD_HEADER = "fremgangsmåde"
# Recipe sheets are tiny; cap the scan so a sheet with phantom dimensions
# (huge declared row count, mostly empty) can't blow up parsing.
MAX_SCAN_ROWS = 120

# Bump when the parsed entry shape, cell source, or URL format changes.
# Cached entries whose ``_v`` differs are re-parsed on the next read so
# callers don't have to know about cache invalidation.
# v2: dish name moved A1 -> C1.
# v3: always use Sheets-API #gid links (even for .xlsx files in Drive), so
#     per-dish deep links work instead of every link going to the file's
#     default tab.
# v4: fall back to openpyxl only when Sheets API returns zero sheets in
#     metadata (signal that developerKey auth can't open the file), not
#     when it returns sheets that just don't match the Ma1/Ti2 pattern.
# v5: drop the bogus ?range= "deep link" from the xlsx fallback — Google
#     Sheets ignores it for tab navigation, so we now leave url="" and the
#     frontend renders a single "Åbn opskriftsark" button instead.
# v6: parse full recipe content (ingredients + Fremgangsmåde steps + weekday)
#     so the app can render recipes in-app instead of only deep-linking.
CACHE_SCHEMA_VERSION = 6


def _grid_cell(grid: list, row: int, col: int):
    """1-indexed cell access into a list-of-rows grid; None if out of range.

    Works for both openpyxl ``iter_rows(values_only=True)`` tuples and the
    Sheets API ``valueRange['values']`` lists (which trim trailing empties).
    """
    if 0 <= row - 1 < len(grid):
        cells = grid[row - 1]
        if cells is not None and 0 <= col - 1 < len(cells):
            return cells[col - 1]
    return None


def _grid_str(grid: list, row: int, col: int) -> str:
    """Trimmed string value of a grid cell ('' if empty/missing)."""
    value = _grid_cell(grid, row, col)
    return "" if value is None else str(value).strip()


def _fmt_amount(value) -> str:
    """Format a quantity for display: trim float noise, keep non-numeric as-is.

    openpyxl returns floats (formula results); the Sheets API returns formatted
    strings that may use a Danish decimal comma. Both collapse to e.g. "7.2",
    "784", "0.45"; anything non-numeric (ranges like "1-2") passes through.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{round(float(value), 2):g}"
    text = str(value).strip()
    try:
        return f"{round(float(text.replace(',', '.')), 2):g}"
    except ValueError:
        return text


def folder_url(folder_id: str) -> str:
    """Return a browser URL for a Google Drive folder."""
    return f"https://drive.google.com/drive/folders/{folder_id}"


class RecipeSheetService:
    """Service for fetching recipe spreadsheets from Google Drive."""

    def __init__(self) -> None:
        self.api_key = settings.GOOGLE_DRIVE_API_KEY
        self.folder_id = settings.GOOGLE_DRIVE_MENU_FOLDER_ID
        self._service = None

    @property
    def service(self):
        """Lazy-load the Google Drive (v3) service."""
        if self._service is None:
            if not self.api_key:
                raise ValueError(
                    "GOOGLE_DRIVE_API_KEY is not configured. "
                    "Please set it in your environment variables."
                )
            self._service = build("drive", "v3", developerKey=self.api_key)
        return self._service

    def _find_spreadsheet_in_folder(self, folder_id: str) -> tuple[str | None, str | None]:
        """Find the first spreadsheet (native Google Sheet or .xlsx) in a folder.

        Returns:
            (file_id, mime_type) or (None, None) if none found.
        """
        try:
            results = (
                self.service.files()
                .list(
                    q=(
                        f"'{folder_id}' in parents and ("
                        f"mimeType='{GSHEET_MIME}' or mimeType='{XLSX_MIME}')"
                    ),
                    fields="files(id, name, mimeType)",
                )
                .execute()
            )
            files = results.get("files", [])
            if not files:
                return None, None
            first = files[0]
            return first["id"], first["mimeType"]
        except Exception as e:
            logger.error(f"Error finding spreadsheet in folder {folder_id}: {e}")
            return None, None

    def parse_recipe_sheets(self, folder_id: str) -> list[dict]:
        """Parse all recipe sheets in the spreadsheet inside ``folder_id``.

        Returns a list of {code, day (0-3), index (int), name (str), url (str)}
        dicts, one per worksheet whose title matches the Ma/Ti/On/To<digit>
        pattern. Returns [] on any failure.
        """
        file_id, _ = self._find_spreadsheet_in_folder(folder_id)
        if not file_id:
            logger.warning(f"No recipe spreadsheet found in folder {folder_id}")
            return []
        return self._parse_with_file_id(folder_id, file_id)

    def _parse_with_file_id(self, folder_id: str, file_id: str) -> list[dict]:
        """Parse the recipe sheets given an already-located spreadsheet file_id.

        Always try the Sheets API path first: it works for both native Google
        Sheets AND .xlsx files in Drive (Drive surfaces them under the same
        ``/spreadsheets/d/<id>/edit`` URL where ``#gid=<sheetId>`` anchors are
        honoured). Fall back to openpyxl only when the Sheets API yields no
        usable metadata (the case when developerKey auth can't open a
        particular .xlsx) — when it returns metadata that just doesn't match
        the Ma1/Ti2 pattern, that's a real "no recipes" answer and we keep it.
        """
        try:
            entries = self._parse_native_sheet(file_id)
        except Exception as e:
            logger.warning(
                "Sheets API parse raised for %s (folder %s), falling back to xlsx: %s",
                file_id,
                folder_id,
                e,
            )
            entries = None

        if entries is not None:
            logger.info(
                "Recipe sheets parsed via Sheets API for %s: %d entries", file_id, len(entries)
            )
            return entries

        logger.info(
            "Sheets API yielded no metadata for %s — using openpyxl xlsx fallback "
            "(per-dish gid links will not be available)",
            file_id,
        )
        try:
            return self._parse_xlsx(file_id)
        except Exception as e:
            logger.error(f"Error parsing recipe sheets in folder {folder_id}: {e}")
            return []

    def _build_entry_from_grid(self, title: str, grid: list, url: str) -> dict | None:
        """Build a full recipe entry from a sheet's cell grid.

        ``grid`` is a list-of-rows (openpyxl tuples or Sheets API value lists).
        Extracts dish name (C1), weekday (F1), the ingredient table (C–F) and
        the Fremgangsmåde steps. Returns None if the title doesn't match the
        recipe-sheet pattern.
        """
        match = SHEET_PATTERN.match(title.strip())
        if not match:
            return None
        prefix = match.group(1).lower()
        index = int(match.group(2))
        day = DAY_PREFIX_TO_INDEX[prefix]

        name = _grid_str(grid, DISH_NAME_ROW, DISH_NAME_COL) or title.strip()
        weekday = DANISH_WEEKDAYS[day]

        # Locate the ingredient-table header ("Ingrediens" in col E) and the
        # "Fremgangsmåde" header so we can split the sheet into its two parts.
        n_rows = min(len(grid), MAX_SCAN_ROWS)
        header_row: int | None = None
        method_row: int | None = None
        for r in range(1, n_rows + 1):
            for c in (COL_STEP, COL_UNIT, COL_INGREDIENT, COL_COMMENT):
                cell = _grid_str(grid, r, c).lower()
                if cell == INGREDIENT_HEADER and header_row is None:
                    header_row = r
                elif cell == METHOD_HEADER and method_row is None:
                    method_row = r

        ing_start = (header_row or 4) + 1
        ing_end = (method_row - 1) if method_row else n_rows

        # Pick the amount column: prefer C, but fall back to B/A for the
        # occasional sheet that only fills an alternate scaling column.
        amount_col = AMOUNT_COL_PREFERENCE[0]
        for candidate in AMOUNT_COL_PREFERENCE:
            if any(
                _grid_cell(grid, r, candidate) not in (None, "")
                for r in range(ing_start, ing_end + 1)
            ):
                amount_col = candidate
                break

        ingredients: list[dict] = []
        for r in range(ing_start, ing_end + 1):
            ing_name = _grid_str(grid, r, COL_INGREDIENT)
            if not ing_name:
                continue
            ingredients.append(
                {
                    "amount": _fmt_amount(_grid_cell(grid, r, amount_col)),
                    "unit": _grid_str(grid, r, COL_UNIT),
                    "name": ing_name,
                    "comment": _grid_str(grid, r, COL_COMMENT),
                }
            )

        steps: list[str] = []
        if method_row:
            for r in range(method_row + 1, n_rows + 1):
                text = _grid_str(grid, r, COL_STEP)
                if text:
                    steps.append(text)

        return {
            "code": title.strip(),
            "day": day,
            "index": index,
            "name": name,
            "weekday": weekday,
            "url": url,
            "ingredients": ingredients,
            "steps": steps,
            "_v": CACHE_SCHEMA_VERSION,
        }

    def _parse_native_sheet(self, file_id: str) -> list[dict] | None:
        """Parse via the Sheets API.

        Works for both native Google Sheets and .xlsx files in Drive — Drive
        exposes the latter under the same ``/spreadsheets/d/<id>/edit`` URL
        and the sheetId returned here is the gid that the URL ``#gid=`` anchor
        expects. Raises on API failure so the caller can fall back to openpyxl.

        Returns:
          * ``None`` if the API returned no sheet metadata at all (caller
            should try the openpyxl fallback);
          * a list (possibly empty) if metadata came back — empty just means
            the spreadsheet has no sheets matching the Ma/Ti/On/To<digit>
            pattern, which is a real answer, not a fallback signal.
        """
        sheets_service = build("sheets", "v4", developerKey=self.api_key)

        # First: titles + gids of all sheets.
        meta = (
            sheets_service.spreadsheets()
            .get(
                spreadsheetId=file_id,
                fields="sheets.properties(sheetId,title)",
            )
            .execute()
        )
        sheets = meta.get("sheets", [])
        if not sheets:
            # No metadata at all — Sheets API can't open this file via
            # developerKey auth (most often the case for .xlsx in Drive).
            return None

        # Collect titles that match, mapping title -> (sheetId, index)
        matching: list[tuple[str, int]] = []
        for sheet in sheets:
            props = sheet.get("properties", {})
            title = props.get("title", "")
            sheet_id = props.get("sheetId", 0)
            if SHEET_PATTERN.match(title.strip()):
                matching.append((title, sheet_id))

        if not matching:
            return []

        # Batch-read columns A–F of each matching sheet (the full recipe block:
        # dish name, ingredient table and Fremgangsmåde steps all live here).
        ranges = [f"'{title}'!A1:F{MAX_SCAN_ROWS}" for title, _ in matching]
        batch = (
            sheets_service.spreadsheets()
            .values()
            .batchGet(spreadsheetId=file_id, ranges=ranges)
            .execute()
        )
        value_ranges = batch.get("valueRanges", [])

        entries: list[dict] = []
        for (title, sheet_id), value_range in zip(matching, value_ranges, strict=False):
            grid = value_range.get("values", [])
            url = f"https://docs.google.com/spreadsheets/d/{file_id}/edit#gid={sheet_id}"
            entry = self._build_entry_from_grid(title, grid, url)
            if entry:
                entries.append(entry)
        return entries

    def _parse_xlsx(self, file_id: str) -> list[dict]:
        """Parse an uploaded .xlsx spreadsheet via openpyxl.

        This is the path that actually runs in production: the menu workbooks
        are Office .xlsx files, which the Sheets API refuses to open, so we
        download and read them locally. ``data_only=True`` returns the cached
        formula results (the quantities are formulas) rather than the formulas.

        We can't get Google's per-sheet gid for an .xlsx via developerKey-auth,
        so entries carry an empty ``url`` and the app renders the parsed recipe
        content in-app instead of deep-linking into a specific tab.
        """
        try:
            import openpyxl

            request = self.service.files().get_media(fileId=file_id)
            file_content = io.BytesIO()
            downloader = MediaIoBaseDownload(file_content, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            file_content.seek(0)

            workbook = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            entries: list[dict] = []
            for ws in workbook.worksheets:
                title = ws.title
                if not SHEET_PATTERN.match(title.strip()):
                    continue
                grid = list(ws.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, values_only=True))
                entry = self._build_entry_from_grid(title, grid, url="")
                if entry:
                    entries.append(entry)
            workbook.close()
            return entries
        except Exception as e:
            logger.error(f"Error parsing xlsx spreadsheet {file_id}: {e}")
            return []

    def get_recipes_for_week(
        self, week_number: int, year: int | None = None, force_refresh: bool = False
    ) -> list[dict]:
        """Get recipe sheets for a week, using/refreshing the DriveMenuCache.

        Returns a list of recipe entry dicts (possibly empty). Cache entries
        whose ``_v`` doesn't match CACHE_SCHEMA_VERSION are treated as stale
        and re-parsed automatically.
        """
        from apps.food.models import DriveMenuCache

        if year is None:
            year = datetime.date.today().year

        cache = DriveMenuCache.objects.filter(week_number=week_number, year=year).first()
        if cache is None:
            return []

        if (
            cache.recipe_sheets
            and not force_refresh
            and self._cache_is_current(cache.recipe_sheets)
        ):
            return cache.recipe_sheets

        if not cache.drive_folder_id:
            return []

        file_id, _ = self._find_spreadsheet_in_folder(cache.drive_folder_id)
        if not file_id:
            return []
        parsed = self._parse_with_file_id(cache.drive_folder_id, file_id)

        cache.recipe_sheets = parsed
        cache.recipe_file_id = file_id or ""
        cache.save(update_fields=["recipe_sheets", "recipe_file_id"])
        return parsed

    @staticmethod
    def _cache_is_current(entries: list[dict]) -> bool:
        """True if every cached entry carries the current schema version."""
        return all(isinstance(e, dict) and e.get("_v") == CACHE_SCHEMA_VERSION for e in entries)

    def recipes_for_date(self, d: datetime.date) -> list[dict]:
        """Return recipe entries for a specific date, sorted by dish index."""
        iso = d.isocalendar()
        week_number = iso[1]
        year = iso[0]
        recipes = self.get_recipes_for_week(week_number, year)
        weekday = d.weekday()
        filtered = [r for r in recipes if r.get("day") == weekday]
        return sorted(filtered, key=lambda r: r.get("index", 0))
